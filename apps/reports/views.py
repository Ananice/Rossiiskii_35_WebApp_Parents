import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from apps.communications.models import Message
from openpyxl.cell.cell import MergedCell

@login_required(login_url='core:login')
def reports_index(request):
    """
    ENDPOINT: GET /reports/
    ОПИСАНИЕ: Страница отчётов — список доступных выгрузок.
    Закрывает требование: личный кабинет (доп. страница на роль).
    """
    from django.shortcuts import render
    context = {
        'breadcrumbs': [
            {'title': 'Главная', 'url': '/', 'icon': 'fas fa-home'},
            {'title': 'Отчёты', 'url': None, 'icon': 'fas fa-file-excel'},
        ]
    }
    return render(request, 'reports/index.html', context)


@login_required(login_url='core:login')
def export_messages_xlsx(request):
    """
    ENDPOINT: GET /reports/export/messages/xlsx/
    ОПИСАНИЕ: Генерация Excel-отчёта по сообщениям пользователя.
    Закрывает требование №14: генерация .xlsx.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Сообщения"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Отчёт по сообщениям — {request.user.get_full_name() or request.user.username}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    headers = ["№", "Дата", "От кого", "Кому", "Текст сообщения", "Прочитано"]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    from django.db.models import Q
    messages = Message.objects.filter(
        Q(sender=request.user) | Q(recipient=request.user)
    ).select_related("sender", "recipient").order_by("-created_at")[:200]

    for i, msg in enumerate(messages, start=1):
        ws.append([
            i,
            msg.created_at.strftime("%d.%m.%Y %H:%M"),
            msg.sender.get_full_name() or msg.sender.username,
            msg.recipient.get_full_name() or msg.recipient.username,
            (msg.content[:80] + "...") if len(msg.content) > 80 else msg.content,
            "Да" if msg.is_read else "Нет",
        ])

    for col in ws.columns:
        col_letter = None
        max_len = 10
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws2 = wb.create_sheet("Информация")
    ws2.append(["Параметр", "Значение"])
    ws2.append(["Пользователь", request.user.get_full_name() or request.user.username])
    ws2.append(["Роль", getattr(request.user, "role", "—")])
    ws2.append(["Дата генерации", timezone.now().strftime("%d.%m.%Y %H:%M")])
    ws2.append(["Всего записей", messages.count()])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"report_messages_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
